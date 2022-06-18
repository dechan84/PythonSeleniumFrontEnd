import openpyxl

class HomePageData:

    test_HomePage_Data = [{"firstname":"Victor", "email":"aa@gmail.com", "gender":"Male"}, {"firstname":"Karen", "email":"ff@gmail.com", "gender":"Female"}]

    # Get values from exel
    # We can declare a static method so we dont need to create an object of this class when we want to use getTestData
    # to do this, we need to remove the self in the method and use @staticmethod
    @staticmethod
    def getTestData(test_case_name):
        # Define a dictionary
        Dict = {}
        # Get the path of the exel file
        book = openpyxl.load_workbook(
            "C:\\Users\\victo\\Documents\\PythonProjectsPyCharm\\PythonSelfFramework\\TestData\\PythonDemo.xlsx")
        # We needs to active the exel file
        sheet = book.active
        for i in range(1, sheet.max_row + 1):
            if sheet.cell(row=i, column=1).value == test_case_name:
                # Start adding values from Testcase2
                for j in range(2, sheet.max_column + 1):
                    # Loop that uses row set to 1 to get all the keys and assign the values to each key
                    Dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value
        print(Dict)
        # We need to send back the parameters in a list!
        return [Dict]