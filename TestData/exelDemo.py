# Example of loading inputs from exel inputs

import openpyxl
# Get the path of the exel file
book = openpyxl.load_workbook("C:\\Users\\victo\\Documents\\PythonProjectsPyCharm\\PythonSelfFramework\\TestData\\PythonDemo.xlsx")
# We needs to active the exel file
sheet = book.active
# Define a dictionary
Dict = {}
# Assign the cel attr to 'cell'
cell = sheet.cell(row=1, column=2)
# Lets print the cell value
print(cell.value)
# We want to write to a cell value now
sheet.cell(row=2, column=2).value = "Victor"
print(sheet.cell(row=2, column=2).value)
print(sheet.max_row)
print(sheet.max_column)

print(sheet['A5'].value)
# Loop to look into a define ranged values in exel, look for Testcase2
# for i in range(1, sheet.max_row+1):
#     if sheet.cell(row=i, column=1).value == "Testcase2":
#         # Start from col 2 and remove test name
#         for j in range (2, sheet.max_column+1):
#             print(sheet.cell(row=i, column=j).value)
# Loop to add values into Dict
for i in range(1, sheet.max_row+1):
    if sheet.cell(row=i, column=1).value == "Testcase2":
        # Start adding values from Testcase2
        for j in range (2, sheet.max_column+1):
            # Loop that uses row set to 1 to get all the keys and assign the values to each key
            Dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value
print(Dict)

